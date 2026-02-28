"""
Script 2: Extract Timestamps from BMP Filenames
================================================
Reads BMP filenames from each trial folder, extracts timestamps embedded in the
filename (format: YYYYMMDDHHMMSS), and calculates elapsed minutes from the start
of each trial. Results are saved to a single Excel workbook with one sheet per trial.

If a filename does not contain a recognisable timestamp, the script falls back to
the file's modification time.

USAGE
-----
1. Edit TRIAL_FOLDERS and OUTPUT_FILE below.
2. Run: python 2_extract_timestamps.py

Dependencies: openpyxl (install with: pip install openpyxl)
"""

import os
import re
from datetime import datetime
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

# =============================================================================
# CONFIGURATION — edit these paths before running
# =============================================================================

TRIAL_FOLDERS = [
    r"D:\All files for santomea and Yakuba\trial 1-santomea control 22-22",
    r"D:\All files for santomea and Yakuba\trial 2-santomea 12-21",
    # Add more folders here, one per line
]

OUTPUT_FILE = r"D:\All files for santomea and Yakuba\timestamps.xlsx"

# =============================================================================
# PROCESSING — do not edit below this line
# =============================================================================

TIMESTAMP_PATTERN = re.compile(r'(\d{14})')  # 14-digit: YYYYMMDDHHMMSS

HEADER = ["File", "Timestamp", "Time_Minutes_From_Start"]
HEADER_FILL = PatternFill("solid", start_color="4F81BD")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def extract_timestamp(filepath):
    filename = os.path.basename(filepath)
    match = TIMESTAMP_PATTERN.search(filename)
    if match:
        try:
            return datetime.strptime(match.group(1), "%Y%m%d%H%M%S")
        except ValueError:
            pass
    return datetime.fromtimestamp(os.path.getmtime(filepath))


def process_folder(folder_path):
    bmp_files = sorted(
        f for f in os.listdir(folder_path) if f.lower().endswith('.bmp')
    )
    if not bmp_files:
        return []

    rows = []
    timestamps = [extract_timestamp(os.path.join(folder_path, f)) for f in bmp_files]
    start = timestamps[0]

    for filename, ts in zip(bmp_files, timestamps):
        elapsed = (ts - start).total_seconds() / 60
        rows.append([filename, ts.strftime("%Y-%m-%d %H:%M:%S"), round(elapsed, 4)])

    return rows


wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

for folder in TRIAL_FOLDERS:
    sheet_name = os.path.basename(folder)[:31]  # Excel sheet name limit
    print(f"\nProcessing: {folder}")

    if not os.path.isdir(folder):
        print(f"  [ERROR] Folder not found.")
        continue

    rows = process_folder(folder)
    if not rows:
        print(f"  [WARNING] No BMP files found.")
        continue

    ws = wb.create_sheet(title=sheet_name)

    for col_idx, header in enumerate(HEADER, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 25

    print(f"  {len(rows)} BMP files processed.")

wb.save(OUTPUT_FILE)
print(f"\nSaved to: {OUTPUT_FILE}")
