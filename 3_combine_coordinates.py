"""
Script 3: Combine YOLO Detection Coordinates
=============================================
Searches each trial folder for YOLO output .txt files, extracts the detection
coordinates (bounding box centre X, centre Y, width, height), and writes all
results to a single Excel workbook with one sheet per trial.

Expected .txt file format (YOLO standard):
    <class> <x_center> <y_center> <width> <height>

USAGE
-----
1. Edit TRIAL_FOLDERS and OUTPUT_FILE below.
2. Run: python 3_combine_coordinates.py

Dependencies: openpyxl (install with: pip install openpyxl)
"""

import os
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

# =============================================================================
# CONFIGURATION — edit these paths before running
# =============================================================================

TRIAL_FOLDERS = [
    r"D:\All files for santomea and Yakuba\trial 1-santomea control 22-22",
    r"D:\All files for santomea and Yakuba\trial 2-santomea 12-21",
    # Add more folders here, one per line
]

OUTPUT_FILE = r"D:\All files for santomea and Yakuba\coordinates.xlsx"

# =============================================================================
# PROCESSING — do not edit below this line
# =============================================================================

HEADER = ["File", "X-center", "Y-center", "Width", "Height"]
HEADER_FILL = PatternFill("solid", start_color="4F81BD")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def parse_txt_file(filepath):
    detections = []
    with open(filepath, 'r') as f:
        for line in f:
            parts = line.strip().split()
            if len(parts) >= 5:
                try:
                    detections.append([float(p) for p in parts[1:5]])
                except ValueError:
                    pass
    return detections


def process_folder(folder_path):
    txt_files = sorted(f for f in os.listdir(folder_path) if f.lower().endswith('.txt'))
    rows = []
    for filename in txt_files:
        filepath = os.path.join(folder_path, filename)
        detections = parse_txt_file(filepath)
        for det in detections:
            rows.append([filename] + det)
        if not detections:
            rows.append([filename, None, None, None, None])
    return rows


wb = openpyxl.Workbook()
wb.remove(wb.active)

for folder in TRIAL_FOLDERS:
    sheet_name = os.path.basename(folder)[:31]
    print(f"\nProcessing: {folder}")

    if not os.path.isdir(folder):
        print(f"  [ERROR] Folder not found.")
        continue

    rows = process_folder(folder)
    if not rows:
        print(f"  [WARNING] No .txt files found.")
        continue

    ws = wb.create_sheet(title=sheet_name)

    for col_idx, header in enumerate(HEADER, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.column_dimensions["A"].width = 35
    for col in ["B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 14

    print(f"  {len(rows)} detections written.")

wb.save(OUTPUT_FILE)
print(f"\nSaved to: {OUTPUT_FILE}")
